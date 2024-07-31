import ast
import os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

class TryExceptVisitor(ast.NodeVisitor):
    def __init__(self):
        self.try_excepts = defaultdict(list)
        self.current_function = None

    def visit_FunctionDef(self, node):
        old_function = self.current_function
        self.current_function = node.name
        self.generic_visit(node)
        self.current_function = old_function

    def visit_Try(self, node):
        func_name = self.current_function or "Global Scope"
        self.try_excepts[func_name].append(node.lineno)
        self.generic_visit(node)

def analyze_file(file_path):
    with open(file_path, 'r') as file:
        tree = ast.parse(file.read())
        visitor = TryExceptVisitor()
        visitor.visit(tree)
        return visitor.try_excepts

def find_python_files(directory, ignore_folders):
    for root, dirs, files in os.walk(directory):
        dirs[:] = [d for d in dirs if d not in ignore_folders]
        for file in files:
            if file.endswith('.py'):
                yield os.path.join(root, file)

def apply_style(cell, is_header=False, is_total=False):
    cell.alignment = Alignment(horizontal='center', vertical='center')
    if is_header:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    elif is_total:
        cell.font = Font(bold=True)

def create_excel_report(data, output_file):
    wb = Workbook()
    
    # Detailed sheet
    ws_detailed = wb.active
    ws_detailed.title = "Detailed Try-Except Summary"

    headers = ["File Path", "Function Name", "Try-Except Count", "Line Numbers"]
    for col, header in enumerate(headers, start=1):
        cell = ws_detailed.cell(row=1, column=col, value=header)
        apply_style(cell, is_header=True)

    row = 2
    for file_path, functions in data.items():
        file_try_except_count = sum(len(lines) for lines in functions.values())
        
        for func_name, lines in functions.items():
            ws_detailed.cell(row=row, column=1, value=file_path)
            ws_detailed.cell(row=row, column=2, value=func_name)
            ws_detailed.cell(row=row, column=3, value=len(lines))
            ws_detailed.cell(row=row, column=4, value=", ".join(map(str, lines)))
            row += 1
        
        # Add a summary row for the file
        for col in range(1, 5):
            cell = ws_detailed.cell(row=row, column=col)
            if col == 1:
                cell.value = file_path
            elif col == 2:
                cell.value = "FILE TOTAL"
            elif col == 3:
                cell.value = file_try_except_count
            apply_style(cell, is_total=True)
        row += 1

    # Summary sheet
    ws_summary = wb.create_sheet(title="Summary")
    
    summary_headers = ["File Path", "Total Try-Except Blocks"]
    for col, header in enumerate(summary_headers, start=1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        apply_style(cell, is_header=True)

    row = 2
    total_blocks = 0
    for file_path, functions in data.items():
        file_try_except_count = sum(len(lines) for lines in functions.values())
        total_blocks += file_try_except_count
        
        ws_summary.cell(row=row, column=1, value=file_path)
        ws_summary.cell(row=row, column=2, value=file_try_except_count)
        row += 1

    # Add total row
    cell = ws_summary.cell(row=row, column=1, value="TOTAL")
    apply_style(cell, is_total=True)
    cell = ws_summary.cell(row=row, column=2, value=total_blocks)
    apply_style(cell, is_total=True)

    # Adjust column widths for both sheets
    for ws in [ws_detailed, ws_summary]:
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

    wb.save(output_file)

def main(directory, ignore_folders):
    data = {}
    total_try_excepts = 0

    for file_path in find_python_files(directory, ignore_folders):
        try_excepts = analyze_file(file_path)
        if try_excepts:
            data[file_path] = try_excepts
            file_try_except_count = sum(len(lines) for lines in try_excepts.values())
            total_try_excepts += file_try_except_count
            print(f"\nFile: {file_path}")
            print(f"Total try-except blocks: {file_try_except_count}")
            for func_name, lines in try_excepts.items():
                print(f"  Function: {func_name}, Count: {len(lines)}")
                print(f"    Lines: {', '.join(map(str, lines))}")
    
    print(f"\nTotal try-except blocks across all files: {total_try_excepts}")

    output_file = "try_except_summary.xlsx"
    create_excel_report(data, output_file)
    print(f"\nExcel report saved as: {output_file}")

if __name__ == "__main__":
    directory = './'
    ignore_folders = []
    main(directory, ignore_folders)